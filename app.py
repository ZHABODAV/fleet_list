```python
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import List
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

st.set_page_config(layout="wide", page_title="Морской Флот - Динамический Планировщик")

# ============================================================
# МОДУЛЬ 1: ПЛЕЧИ И МАРШРУТЫ
# ============================================================

@dataclass
class Leg:
    """Плечо маршрута"""
    name: str
    port_from: str
    port_to: str
    duration_days: float
    operation_days: float

    def total_time(self):
        return self.duration_days + self.operation_days

    def __repr__(self):
        return f"{self.port_from}→{self.port_to}"


@dataclass
class Route:
    """Маршрут = набор плеч"""
    name: str
    legs: List[Leg]
    ship_capacity: float

    def total_time(self):
        return sum(leg.total_time() for leg in self.legs)

    def get_ports(self):
        ports = [self.legs[0].port_from]
        for leg in self.legs:
            ports.append(leg.port_to)
        return ports


class VoyageScheduler:
    """Планировщик рейсов"""

    def __init__(self, route: Route, start_date: datetime, num_ships: int, interval_days: float = 1):
        self.route = route
        self.start_date = start_date
        self.num_ships = num_ships
        self.interval_days = interval_days
        self.voyages = []

        self._generate_voyages()

    def _generate_voyages(self):
        for ship_id in range(1, self.num_ships + 1):
            departure = self.start_date + timedelta(days=(ship_id - 1) * self.interval_days)
            voyage = self._calculate_voyage(ship_id, departure)
            self.voyages.append(voyage)

    def _calculate_voyage(self, ship_id: int, departure: datetime):
        current_time = departure
        voyage = {
            "ship_id": ship_id,
            "route": self.route.name,
            "ship_name": f"Ship_{ship_id}",
            "capacity": self.route.ship_capacity,
            "itinerary": []
        }

        for i, leg in enumerate(self.route.legs):
            leg_start = current_time
            arrival = leg_start + timedelta(days=leg.duration_days)
            leg_end = arrival + timedelta(days=leg.operation_days)

            voyage["itinerary"].append({
                "leg_no": i + 1,
                "leg": leg.name,
                "port_from": leg.port_from,
                "port_to": leg.port_to,
                "departure": leg_start,
                "arrival": arrival,
                "operation_end": leg_end,
                "voyage_time_days": leg.duration_days,
                "operation_days": leg.operation_days
            })

            current_time = leg_end

        voyage["total_voyage_time"] = (current_time - departure).days
        voyage["arrival_final"] = current_time

        return voyage

    def get_schedule_df(self):
        rows = []
        for voyage in self.voyages:
            for stop in voyage["itinerary"]:
                rows.append({
                    "Судно": voyage["ship_name"],
                    "Плечо": stop["leg"],
                    "Из": stop["port_from"],
                    "В": stop["port_to"],
                    "Отправление": stop["departure"].strftime("%Y-%m-%d"),
                    "Прибытие": stop["arrival"].strftime("%Y-%m-%d"),
                    "Конец_операции": stop["operation_end"].strftime("%Y-%m-%d"),
                    "В_пути_дни": stop["voyage_time_days"],
                    "Операция_дни": stop["operation_days"]
                })

        return pd.DataFrame(rows)


class TankSimulation:
    """Симуляция буферной ёмкости"""

    def __init__(self, capacity: float, river_scheduler: VoyageScheduler, sea_scheduler: VoyageScheduler):
        self.capacity = capacity
        self.level = 0
        self.river_scheduler = river_scheduler
        self.sea_scheduler = sea_scheduler
        self.events = []
        self.tank_log = []

        self._simulate()

    def _simulate(self):
        transfer_port = self.river_scheduler.route.get_ports()[-1]

        for voyage in self.river_scheduler.voyages:
            last_stop = voyage["itinerary"][-1]
            self.events.append({
                "time": last_stop["arrival"],
                "type": "river_arrival",
                "ship": voyage["ship_name"],
                "port": transfer_port,
                "volume": voyage["capacity"]
            })

        for voyage in self.sea_scheduler.voyages:
            first_stop = voyage["itinerary"][0]
            self.events.append({
                "time": first_stop["departure"],
                "type": "sea_departure",
                "ship": voyage["ship_name"],
                "port": transfer_port,
                "volume": voyage["capacity"]
            })

        self.events.sort(key=lambda x: x["time"])

        for event in self.events:
            if event["type"] == "river_arrival":
                self.level += event["volume"]
                if self.level > self.capacity:
                    self.level = self.capacity
                self.tank_log.append({
                    "time": event["time"],
                    "ship": event["ship"],
                    "action": f"Выгрузка речного (+{event['volume']})",
                    "tank_level": self.level,
                    "port": event["port"]
                })

            elif event["type"] == "sea_departure":
                self.level -= event["volume"]
                if self.level < 0:
                    self.level = 0
                self.tank_log.append({
                    "time": event["time"],
                    "ship": event["ship"],
                    "action": f"Загрузка морского (-{event['volume']})",
                    "tank_level": self.level,
                    "port": event["port"]
                })

    def get_log_df(self):
        return pd.DataFrame(self.tank_log)


# ============================================================
# МОДУЛЬ 2: ЭКСПОРТ В EXCEL
# ============================================================

def export_to_excel(river_scheduler, sea_scheduler, tank_sim):
    """Экспорт всех данных в Excel"""

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Лист 1: Речной флот
        df_river = river_scheduler.get_schedule_df()
        df_river.to_excel(writer, sheet_name='Речной_флот', index=False)

        # Лист 2: Морской флот
        df_sea = sea_scheduler.get_schedule_df()
        df_sea.to_excel(writer, sheet_name='Морской_флот', index=False)

        # Лист 3: Динамика танка
        df_tank = tank_sim.get_log_df()
        df_tank.to_excel(writer, sheet_name='Динамика_танка', index=False)

        # Лист 4: Итоги
        df_summary = pd.DataFrame({
            "Параметр": [
                "Количество речных судов",
                "Количество морских судов",
                "Kapacитет речного судна",
                "Kapacитет морского судна",
                "Max уровень танка",
                "Min уровень танка",
                "Ёмкость танка",
                "Начало речных рейсов",
                "Начало морских рейсов",
                "Общее время цикла (дни)"
            ],
            "Значение": [
                river_scheduler.num_ships,
                sea_scheduler.num_ships,
                river_scheduler.route.ship_capacity,
                sea_scheduler.route.ship_capacity,
                max([l['tank_level'] for l in tank_sim.tank_log]),
                min([l['tank_level'] for l in tank_sim.tank_log]),
                tank_sim.capacity,
                river_scheduler.start_date.strftime("%Y-%m-%d"),
                sea_scheduler.start_date.strftime("%Y-%m-%d"),
                max([(v['arrival_final'] - river_scheduler.voyages[0]['itinerary'][0]['departure']).days 
                     for v in river_scheduler.voyages + sea_scheduler.voyages])
            ]
        })
        df_summary.to_excel(writer, sheet_name='Итоги', index=False)

    output.seek(0)
    return output


def style_excel(excel_file):
    """Стилизация Excel файла"""
    wb = openpyxl.load_workbook(excel_file)

    # Стиль для заголовков
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Применяем стили к заголовкам
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

        # Применяем стили к остальным ячейкам
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column <= 3 or "Дата" in str(ws.cell(1, cell.column).value):
                    cell.alignment = center_alignment

        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    return wb


# ============================================================
# МОДУЛЬ 3: ЭКСПОРТ В PDF
# ============================================================

def export_to_pdf(river_scheduler, sea_scheduler, tank_sim, 
                  fig_gantt, fig_tank, fig_gantt_detail):
    """Экспорт в PDF с графиками и таблицами"""

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )

    story = []

    # Титул
    story.append(Paragraph("⚓ ОТЧЁТ О ПЛАНИРОВАНИИ МОРСКОЙ ПЕРЕВАЛКИ", title_style))
    story.append(Paragraph(f"Дата отчёта: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 12))

    # Итоги
    story.append(Paragraph("Общие Итоги", heading_style))
    summary_data = [
        ["Параметр", "Значение"],
        ["Речные суда", str(river_scheduler.num_ships)],
        ["Морские суда", str(sea_scheduler.num_ships)],
        ["Капацитет танка", f"{tank_sim.capacity} единиц"],
        ["Max уровень танка", f"{max([l['tank_level'] for l in tank_sim.tank_log])} единиц"],
        ["Min уровень танка", f"{min([l['tank_level'] for l in tank_sim.tank_log])} единиц"],
    ]

    summary_table = Table(summary_data, colWidths=[4*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F0F8')])
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    # График 1: Ганта
    story.append(PageBreak())
    story.append(Paragraph("Диаграмма Ганта", heading_style))

    # Сохраняем граф в временный PNG
    img_gantt_path = "/tmp/gantt.png"
    fig_gantt.write_image(img_gantt_path, width=1200, height=400)
    story.append(Image(img_gantt_path, width=7*inch, height=2.5*inch))
    story.append(Spacer(1, 12))

    # График 2: Танк
    story.append(PageBreak())
    story.append(Paragraph("Динамика Буферной Ёмкости", heading_style))

    img_tank_path = "/tmp/tank.png"
    fig_tank.write_image(img_tank_path, width=1200, height=400)
    story.append(Image(img_tank_path, width=7*inch, height=2.5*inch))
    story.append(Spacer(1, 12))

    # Таблица речного флота
    story.append(PageBreak())
    story.append(Paragraph("Расписание Речного Флота", heading_style))

    df_river = river_scheduler.get_schedule_df()
    river_data = [list(df_river.columns)] + df_river.values.tolist()

    river_table = Table(river_data, colWidths=[1.2*inch, 1.2*inch, 0.8*inch, 0.8*inch, 
                                               1*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
    river_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F0F8')])
    ]))
    story.append(river_table)
    story.append(Spacer(1, 12))

    # Таблица морского флота
    story.append(PageBreak())
    story.append(Paragraph("Расписание Морского Флота", heading_style))

    df_sea = sea_scheduler.get_schedule_df()
    sea_data = [list(df_sea.columns)] + df_sea.values.tolist()

    sea_table = Table(sea_data, colWidths=[1.2*inch, 1.2*inch, 0.8*inch, 0.8*inch,
                                           1*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
    sea_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF8C42')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF0E8')])
    ]))
    story.append(sea_table)
    story.append(Spacer(1, 12))

    # Таблица танка
    story.append(PageBreak())
    story.append(Paragraph("Динамика Буферной Ёмкости", heading_style))

    df_tank = tank_sim.get_log_df()
    df_tank['time'] = df_tank['time'].dt.strftime('%Y-%m-%d %H:%M')
    tank_data = [list(df_tank.columns)] + df_tank.values.tolist()

    tank_table = Table(tank_data, colWidths=[1.5*inch, 1.2*inch, 2.5*inch, 1.2*inch, 1*inch])
    tank_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#006600')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F8E8')])
    ]))
    story.append(tank_table)

    # Генерируем PDF
    doc.build(story)
    output.seek(0)
    return output


# ============================================================
# МОДУЛЬ 4: ИНТЕРАКТИВНЫЕ ГРАФИКИ
# ============================================================

def create_gantt_chart(river_scheduler, sea_scheduler):
    """Интерактивная диаграмма Ганта"""
    fig = go.Figure()

    y_pos = 0
    y_labels = []

    # Речные суда
    for i, voyage in enumerate(river_scheduler.voyages):
        color = f"rgba(100, 150, 255, 0.8)"
        y_labels.append(voyage["ship_name"])

        for stop in voyage["itinerary"]:
            fig.add_trace(go.Bar(
                x=[stop["operation_end"] - stop["departure"]],
                y=[y_pos],
                base=stop["departure"],
                orientation='h',
                name=voyage["ship_name"],
                marker_color=color,
                text=f"{stop['port_from'][-1]}→{stop['port_to'][-1]}",
                textposition="inside",
                hovertemplate=f"<b>{voyage['ship_name']}</b><br>Маршрут: {stop['leg']}<br>Отправление: %{{base}}<br>Окончание: %{{x}}<extra></extra>",
                showlegend=(i == 0)
            ))
        y_pos += 1

    # Морские суда
    for i, voyage in enumerate(sea_scheduler.voyages):
        color = f"rgba(255, 165, 100, 0.8)"
        y_labels.append(voyage["ship_name"])

        for stop in voyage["itinerary"]:
            fig.add_trace(go.Bar(
                x=[stop["operation_end"] - stop["departure"]],
                y=[y_pos],
                base=stop["departure"],
                orientation='h',
                name=voyage["ship_name"],
                marker_color=color,
                text=f"{stop['port_from'][-1]}→{stop['port_to'][-1]}",
                textposition="inside",
                hovertemplate=f"<b>{voyage['ship_name']}</b><br>Маршрут: {stop['leg']}<br>Отправление: %{{base}}<br>Окончание: %{{x}}<extra></extra>",
                showlegend=(i == 0)
            ))
        y_pos += 1

    fig.update_layout(
        title="Диаграмма Ганта - Расписание Флота",
        xaxis_title="Дата",
        yaxis_title="Суда",
        hovermode='closest',
        barmode='overlay',
        height=400 + len(y_labels) * 15,
        yaxis=dict(
            tickmode='linear',
            tick0=0,
            dtick=1,
            ticktext=y_labels,
            tickvals=list(range(len(y_labels)))
        ),
        xaxis={'type': 'date'}
    )

    return fig


def create_tank_chart(tank_sim):
    """Интерактивный график танка"""
    df = tank_sim.get_log_df()

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df['time'],
        y=df['tank_level'],
        mode='lines+markers',
        name='Уровень танка',
        line=dict(color='darkblue', width=3),
        marker=dict(size=8),
        fill='tozeroy',
        fillcolor='rgba(100, 150, 255, 0.3)',
        hovertemplate="<b>%{x|%Y-%m-%d}</b><br>Уровень: %{y} единиц<br>%{customdata}<extra></extra>",
        customdata=df['action']
    ))

    fig.add_hline(y=tank_sim.capacity, line_dash="dash", line_color="red", 
                 annotation_text=f"Макс: {tank_sim.capacity}",
                 annotation_position="right")

    fig.update_layout(
        title="Динамика Буферной Ёмкости",
        xaxis_title="Дата",
        yaxis_title="Уровень (единицы)",
        hovermode='x unified',
        height=400,
        xaxis={'type': 'date'}
    )

    return fig


def create_gantt_detailed(river_scheduler, sea_scheduler):
    """Детальный Ганта с цветом по судам"""
    df_rows = []

    for voyage in river_scheduler.voyages:
        for stop in voyage["itinerary"]:
            df_rows.append({
                "Task": voyage["ship_name"],
                "Start": stop["departure"],
                "Finish": stop["operation_end"],
                "Resource": "River",
                "Resource_num": voyage["ship_id"]
            })

    for voyage in sea_scheduler.voyages:
        for stop in voyage["itinerary"]:
            df_rows.append({
                "Task": voyage["ship_name"],
                "Start": stop["departure"],
                "Finish": stop["operation_end"],
                "Resource": "Sea",
                "Resource_num": voyage["ship_id"]
            })

    df = pd.DataFrame(df_rows)

    fig = px.timeline(
        df,
        x_start="Start",
        x_end="Finish",
        y="Task",
        color="Resource",
        title="Детальное Расписание",
        color_discrete_map={"River": "lightblue", "Sea": "lightcoral"},
        hover_data=["Start", "Finish"],
        height=400 + len(set(df['Task'])) * 15
    )

    fig.update_layout(
        xaxis={'type': 'date'},
        hovermode='closest'
    )

    return fig


# ============================================================
# ГЛАВНОЕ ПРИЛОЖЕНИЕ STREAMLIT
# ============================================================

st.title("⚓ Динамический Планировщик Морской Перевалки")
st.markdown("---")

# БОКОВАЯ ПАНЕЛЬ - ПАРАМЕТРЫ
with st.sidebar:
    st.header("Параметры Симуляции")

    st.subheader("Речной Флот")
    river_start = st.date_input(
        "Начало речных рейсов",
        value=datetime(2024, 4, 10),
        key="river_start"
    )
    river_ships = st.slider(
        "Количество речных судов",
        min_value=1,
        max_value=10,
        value=3,
        key="river_ships"
    )
    river_interval = st.slider(
        "Интервал между судами (дни)",
        min_value=0.5,
        max_value=5.0,
        value=1.0,
        step=0.5,
        key="river_interval"
    )

    st.subheader("Морской Флот")
    sea_start = st.date_input(
        "Начало морских рейсов",
        value=datetime(2024, 11, 20),
        key="sea_start"
    )
    sea_ships = st.slider(
        "Количество морских судов",
        min_value=1,
        max_value=10,
        value=4,
        key="sea_ships"
    )
    sea_interval = st.slider(
        "Интервал между судами (дни)",
        min_value=0.5,
        max_value=5.0,
        value=1.0,
        step=0.5,
        key="sea_interval"
    )

    st.subheader("Маршруты")
    river_duration = st.slider(
        "Речное плечо A→B (дни в пути)",
        min_value=1,
        max_value=20,
        value=5,
        key="river_duration"
    )
    river_op_time = st.slider(
        "Операция в портe B (дни)",
        min_value=0.5,
        max_value=5.0,
        value=1.0,
        step=0.5,
        key="river_op"
    )

    sea_duration = st.slider(
        "Морское плечо B→C (дни в пути)",
        min_value=1,
        max_value=30,
        value=10,
        key="sea_duration"
    )
    sea_op_time = st.slider(
        "Операция в порту C (дни)",
        min_value=0.5,
        max_value=5.0,
        value=1.0,
        step=0.5,
        key="sea_op"
    )

    st.subheader("Ёмкости")
    river_capacity = st.slider(
        "Грузоподъёмность речного судна (единицы)",
        min_value=10,
        max_value=100,
        value=30,
        step=5,
        key="river_cap"
    )
    sea_capacity = st.slider(
        "Грузоподъёмность морского судна (единицы)",
        min_value=10,
        max_value=100,
        value=25,
        step=5,
        key="sea_cap"
    )
    tank_capacity = st.slider(
        "Максимум буферного танка (единицы)",
        min_value=50,
        max_value=500,
        value=100,
        step=10,
        key="tank_cap"
    )


# СОЗДАНИЕ МАРШРУТОВ И РАСПИСАНИЙ
leg_river = Leg(
    name="Речное_A-B",
    port_from="PortA",
    port_to="PortB",
    duration_days=river_duration,
    operation_days=river_op_time
)

leg_sea = Leg(
    name="Морское_B-C",
    port_from="PortB",
    port_to="PortC",
    duration_days=sea_duration,
    operation_days=sea_op_time
)

route_river = Route(
    name="Речной_маршрут",
    legs=[leg_river],
    ship_capacity=river_capacity
)

route_sea = Route(
    name="Морской_маршрут",
    legs=[leg_sea],
    ship_capacity=sea_capacity
)

river_scheduler = VoyageScheduler(
    route=route_river,
    start_date=datetime.combine(river_start, datetime.min.time()),
    num_ships=river_ships,
    interval_days=river_interval
)

sea_scheduler = VoyageScheduler(
    route=route_sea,
    start_date=datetime.combine(sea_start, datetime.min.time()),
    num_ships=sea_ships,
    interval_days=sea_interval
)

tank_sim = TankSimulation(tank_capacity, river_scheduler, sea_scheduler)

# Создаём графики
fig_gantt = create_gantt_chart(river_scheduler, sea_scheduler)
fig_tank = create_tank_chart(tank_sim)
fig_gantt_detail = create_gantt_detailed(river_scheduler, sea_scheduler)

# КНОПКИ ЭКСПОРТА
st.markdown("### Экспорт Данных")

col1, col2, col3 = st.columns(3)

with col1:
    # Excel экспорт
    excel_file = export_to_excel(river_scheduler, sea_scheduler, tank_sim)
    wb = style_excel(excel_file)

    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    st.download_button(
        label="Скачать Excel",
        data=excel_output,
        file_name=f"fleet_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with col2:
    # PDF экспорт
    pdf_file = export_to_pdf(river_scheduler, sea_scheduler, tank_sim, 
                            fig_gantt, fig_tank, fig_gantt_detail)

    st.download_button(
        label="Скачать PDF",
        data=pdf_file,
        file_name=f"fleet_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        mime="application/pdf",
        use_container_width=True
    )

with col3:
    # CSV экспорт
    csv_data = river_scheduler.get_schedule_df().to_csv(index=False)
    st.download_button(
        label="Скачать CSV",
        data=csv_data,
        file_name=f"fleet_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv",
        use_container_width=True
    )

st.markdown("---")

# ОСНОВНЫЕ ВКЛАДКИ
tabs = st.tabs([
    "📊 Диаграмма Ганта",
    "📈 Динамика Танка",
    "⚙️ Детальное Расписание",
    "📋 Таблицы"
])

with tabs[0]:
    st.plotly_chart(fig_gantt, use_container_width=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Всего судов", river_ships + sea_ships)
    with col2:
        total_days = max([(v['arrival_final'] - river_scheduler.voyages[0]['itinerary'][0]['departure']).days 
                         for v in river_scheduler.voyages + sea_scheduler.voyages])
        st.metric("Длительность цикла", f"{total_days} дней")
    with col3:
        st.metric("Танк max/min", f"{max([l['tank_level'] for l in tank_sim.tank_log])}/{min([l['tank_level'] for l in tank_sim.tank_log])}")

with tabs[1]:
    st.plotly_chart(fig_tank, use_container_width=True)

    df_tank = tank_sim.get_log_df()
    st.dataframe(df_tank, use_container_width=True)

with tabs[2]:
    st.plotly_chart(fig_gantt_detail, use_container_width=True)

with tabs[3]:
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Речной Флот")
        st.dataframe(river_scheduler.get_schedule_df(), use_container_width=True)

    with col2:
        st.subheader("Морской Флот")
        st.dataframe(sea_scheduler.get_schedule_df(), use_container_width=True)

st.markdown("---")
st.markdown("**Разработано для оптимизации морских перевозок** ⚓")
``